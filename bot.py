from telegram import Update, KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ConversationHandler, ContextTypes
from datetime import datetime
import jdatetime
import pytz
from openpyxl import Workbook, load_workbook
import os

# مراحل مکالمه
NAME, PHONE, CONFIRM_PHONE, NEW_PHONE = range(4)

# شروع مکالمه
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("سلام! لطفاً نام و نام خانوادگی خود را وارد کنید:")
    return NAME

# دریافت نام
async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['name'] = update.message.text

    button = KeyboardButton("📱 ارسال شماره تماس", request_contact=True)
    reply_markup = ReplyKeyboardMarkup([[button]], resize_keyboard=True, one_time_keyboard=True)

    await update.message.reply_text(
        "لطفاً شماره موبایل خود را با زدن دکمه زیر ارسال کنید:",
        reply_markup=reply_markup
    )
    return PHONE

# دریافت شماره تلگرام یا دستی
async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.contact:
        phone_number = update.message.contact.phone_number
        if phone_number.startswith("98") and len(phone_number) == 12:
            phone_number = "0" + phone_number[2:]
    else:
        phone_number = update.message.text

    context.user_data['phone_candidate'] = phone_number

    await update.message.reply_text(
        f"شماره دریافت شده: {phone_number}\n"
        "آیا می‌خواهید اس‌ام‌اس‌های مربوط به وبینار به همین شماره ارسال شود؟ (بله/خیر)",
        reply_markup=ReplyKeyboardMarkup(
            [["بله", "خیر"]],
            resize_keyboard=True,
            one_time_keyboard=True
        )
    )
    return CONFIRM_PHONE

# تأیید شماره یا گرفتن شماره جدید
async def confirm_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_response = update.message.text.strip().lower()

    if user_response in ["بله", "اوکی", "ok", "yes"]:
        context.user_data['phone'] = context.user_data['phone_candidate']
        return await save_registration(update, context)

    elif user_response in ["خیر", "نه", "no"]:
        await update.message.reply_text(
            "لطفاً شماره صحیح خود را وارد کنید:",
            reply_markup=ReplyKeyboardRemove()
        )
        return NEW_PHONE

    else:
        await update.message.reply_text("لطفاً فقط یکی از گزینه‌ها را انتخاب کنید (بله یا خیر).")
        return CONFIRM_PHONE

# دریافت شماره جدید
async def get_new_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['phone'] = update.message.text
    return await save_registration(update, context)

# ذخیره اطلاعات در اکسل
async def save_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tehran = pytz.timezone("Asia/Tehran")
    now = datetime.now(tehran)

    jalali_date = jdatetime.datetime.fromgregorian(datetime=now)
    date_str = jalali_date.strftime("%Y/%m/%d")
    time_str = jalali_date.strftime("%H:%M")

    file_name = "webinar_signups.xlsx"

    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "ثبت‌نام‌ها"
        ws.append(["نام و نام خانوادگی", "شماره تماس", "تاریخ ثبت‌نام (شمسی)", "ساعت ثبت‌نام"])
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    ws.append([context.user_data['name'], context.user_data['phone'], date_str, time_str])
    wb.save(file_name)

    await update.message.reply_text(
        "ثبت‌نام شما با موفقیت انجام شد ✅\n"
        "📌 اطلاعات وبینار:\n"
        "🎤 وبینار فن بیان و اعتماد به نفس\n"
        "📅 دوشنبه ۸ اردیبهشت\n"
        "🕕 ساعت ۱۸:۰۰\n"
        "📍 به‌زودی لینک برایتان ارسال می‌شود 🌟",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

# لغو عملیات
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("عملیات لغو شد.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# اجرای بات
def main():
    app = Application.builder().token("7081860324:AAE64TaHSlfNsNXZ6SDP1zoSgTawekaUSNo").build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            PHONE: [MessageHandler(filters.CONTACT | (filters.TEXT & ~filters.COMMAND), get_phone)],
            CONFIRM_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_phone)],
            NEW_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_new_phone)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv_handler)
    app.run_polling()

if __name__ == '__main__':
    main()
